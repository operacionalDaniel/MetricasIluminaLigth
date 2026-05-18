#!/usr/bin/env python3
# ─────────────────────────────────────────────────────────────────────
# processar_operacional.py
# Lê as 3 planilhas exportadas do sistema e gera dados_operacional.json
#
# USO:
#   python processar_operacional.py ocorrencias.xlsx materiais.xlsx prazos.xlsx
#   python processar_operacional.py  (usa nomes padrão)
# ─────────────────────────────────────────────────────────────────────
import sys, os, json, re, subprocess
from collections import Counter
from datetime import datetime

try:
    import openpyxl
except ImportError:
    subprocess.run([sys.executable,"-m","pip","install","openpyxl","--break-system-packages","-q"])
    import openpyxl

# ── ARQUIVOS (argumento ou nome padrão) ──────────────────────────────
def find_file(keywords, default):
    if os.path.exists(default): return default
    for f in os.listdir('.'):
        if f.endswith('.xlsx') and any(k.lower() in f.lower() for k in keywords):
            return f
    return None

args = sys.argv[1:]
FILE_OC  = args[0] if len(args)>0 else find_file(['ocorrencia','atendimento','sheet'], 'ocorrencias.xlsx')
FILE_MAT = args[1] if len(args)>1 else find_file(['material','materiais'], 'materiais.xlsx')
FILE_PR  = args[2] if len(args)>2 else find_file(['prazo','relatorio'], 'prazos.xlsx')

print(f"\n📂 Arquivos:")
print(f"   Ocorrências : {FILE_OC}")
print(f"   Materiais   : {FILE_MAT}")
print(f"   Prazos      : {FILE_PR}")

def load(path):
    if not path or not os.path.exists(path):
        print(f"   ⚠  Não encontrado: {path}")
        return None
    return openpyxl.load_workbook(path, data_only=True)

# ── OCORRÊNCIAS ──────────────────────────────────────────────────────
result = {}
mes_str = datetime.now().strftime('%b/%y').lower()

wb_oc = load(FILE_OC)
if wb_oc:
    ws = wb_oc.active
    rows = list(ws.iter_rows(values_only=True))
    hdrs = [str(v or '').strip() for v in rows[0]]
    ci = {h:i for i,h in enumerate(hdrs)}

    status_c  = Counter()
    origem_c  = Counter()
    equipe_c  = Counter()
    motivo_c  = Counter()
    solucao_c = Counter()
    bairro_c  = Counter()
    tipo_oc_c = Counter()
    prazo_c   = Counter()
    tempos    = []

    for row in rows[1:]:
        if not any(row): continue
        def g(field, default=''):
            idx = ci.get(field, -1)
            return str(row[idx] or default).strip() if idx >= 0 else default

        status_c [g('desc_status_atendimento_ps')] += 1
        origem_c [g('desc_tipo_origem_ocorrencia')] += 1
        equipe_c [g('desc_equipe')]                 += 1
        motivo_c [g('desc_motivo_atendimento_ps')]  += 1
        solucao_c[g('desc_solucao_atendimento_ps')] += 1
        bairro_c [g('nome_bairro')]                 += 1
        tipo_oc_c[g('desc_tipo_ocorrencia')]        += 1
        prazo_c  [g('desc_prazo')]                  += 1
        try:
            t = float(str(row[ci.get('prazo',-1)] or 0).replace(',','.'))
            if 0 < t < 500: tempos.append(t)
        except: pass

    no_prazo   = sum(v for k,v in prazo_c.items() if 'no prazo' in k.lower())
    fora_prazo = sum(v for k,v in prazo_c.items() if 'atraso' in k.lower())

    # Prazo por origem — cruzamento com arquivo de prazos (se disponível)
    # Estes valores são contados pelo arquivo B (prazos)
    # Mapeados diretamente do arquivo C (ocorrências) via campo desc_prazo por enquanto

    # Detecta mês do arquivo
    for row in rows[1:10]:
        for cell in row:
            m = re.search(r'(\d{4}-(\d{2})-\d{2})', str(cell or ''))
            if m:
                mn = int(m.group(2))
                meses_pt = ['jan','fev','mar','abr','mai','jun','jul','ago','set','out','nov','dez']
                mes_str = f"{meses_pt[mn-1]}/26"
                break

    total = sum(status_c.values())
    bairros = bairro_c.most_common(10)

    result = {
        "mes": mes_str,
        "total_os": total,
        "atendidas": status_c.get('Atendido', 0),
        "modernizacao": status_c.get('Modernização', 0),
        "encontrado_normal": status_c.get('Encontrado normal', 0),
        "fora_escopo": status_c.get('Fora do escopo do contrato', 0),
        "projeto": status_c.get('Projeto', 0),
        "outro_atendimento": status_c.get('Outro Atendimento', 0),
        "impossibilidade": status_c.get('Impossibilidade', 0),
        "repetido": status_c.get('Repetido', 0),
        "nova_instalacao": status_c.get('Nova Instalação', 0),
        "no_prazo": no_prazo,
        "fora_prazo": fora_prazo,
        "prazo_ronda_ok": 0,     # preencher ao cruzar com arquivo de prazos
        "prazo_ronda_fora": 0,
        "prazo_cc_ok": 0,
        "prazo_chatbot_ok": 0,
        "prazo_app_ok": 0,
        "prazo_telegestao_ok": 0,
        "prazo_telegestao_fora": 0,
        "prazo_contratual_h": 48,
        "tempo_medio_h": round(sum(tempos)/len(tempos),1) if tempos else 0,
        "equipe01": equipe_c.get('Equipe 01', 0),
        "equipe02": equipe_c.get('Equipe 02', 0),
        "origem_ronda": origem_c.get('Ronda própria', 0),
        "origem_callcenter": origem_c.get('Call Center', 0),
        "origem_chatbot": origem_c.get('Chatbot', 0),
        "origem_app": origem_c.get('App', 0),
        "origem_telegestao": sum(v for k,v in origem_c.items()
                                 if any(x in k.lower() for x in ['tele','telege','teleger','gestão'])),
        "motivo_rele": motivo_c.get('Rele queimado', 0),
        "motivo_lampada": motivo_c.get('Lâmpada queimada', 0),
        "motivo_conexao": motivo_c.get('Conexão', 0),
        "motivo_luminaria": motivo_c.get('Luminária LED 30W', 0),
        "motivo_particular": motivo_c.get('Propriedade Particular', 0),
        "motivo_chuva": motivo_c.get('Chuva', 0),
        "motivo_outros": sum(v for k,v in motivo_c.items()
                            if k not in ('Rele queimado','Lâmpada queimada','Conexão',
                                         'Luminária LED 30W','Propriedade Particular','Chuva','')),
        "solucao_troca": solucao_c.get('Troca itens', 0),
        "solucao_manutencao": solucao_c.get('Manutenção', 0),
        "solucao_sem_acao": solucao_c.get('', 0),
        "tipo_avulso": tipo_oc_c.get('Atendimento Avulso', 0),
        "tipo_lampada_apagada": tipo_oc_c.get('Lâmpada apagada', 0),
        "tipo_outros": sum(v for k,v in tipo_oc_c.items()
                          if k not in ('Atendimento Avulso','Lâmpada apagada','')),
    }
    # Bairros top 10
    for rank, (bairro, qtd) in enumerate(bairros[:10], 1):
        result[f'bairro_{rank}'] = bairro
        result[f'bairro_{rank}_qtd'] = qtd

    print(f"\n✅ Ocorrências: {total} OS · {mes_str}")

# ── MATERIAIS ────────────────────────────────────────────────────────
wb_mat = load(FILE_MAT)
mat_data = {}
if wb_mat:
    ws = wb_mat.active
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[6:]:
        if row[0] and str(row[0]) not in ('Material','None') and row[-1]:
            try:
                mat_data[str(row[0]).strip()] = int(float(row[-1]))
            except: pass

    def mq(keys):
        for k,v in mat_data.items():
            if any(key.lower() in k.lower() for key in keys): return v
        return 0

    result.update({
        "mat_rele_sgip":     mq(['SGIP','SMARTGREEN']),
        "mat_led30w_reeme":  mq(['LED 30W','30W - REEME']),
        "mat_conector2":     mq(['tipo II','cunha de BT, tipo II']),
        "mat_led40w_reeme":  mq(['LED 40W','40W - REEME']),
        "mat_led120w_reeme": mq(['LED 120W','120W - REEME']),
        "mat_braco2m":       mq(['Braço 2','braço 2']),
        "mat_conector1":     mq(['tipo I (cinza)']),
        "mat_led70w_reeme":  mq(['LED 70W','70W - REEME']),
        "mat_argos30w":      mq(['ARGOS']),
        "mat_outros":        sum(v for k,v in mat_data.items()
                                if not any(x in k.upper() for x in ['SGIP','LED','ARGOS','BRAÇO','BRACO']))
    })
    print(f"✅ Materiais: {len(mat_data)} tipos")

# ── SALVAR JSON ──────────────────────────────────────────────────────
JSON_FILE = 'dados_operacional.json'
try:
    with open(JSON_FILE, 'r', encoding='utf-8') as f:
        existing = json.load(f)
except:
    existing = {"ultima_atualizacao": "", "meses": {}}

existing['ultima_atualizacao'] = datetime.now().strftime('%d/%m/%Y %H:%M')
existing['meses'][mes_str] = result

with open(JSON_FILE, 'w', encoding='utf-8') as f:
    json.dump(existing, f, ensure_ascii=False, indent=2)

print(f"\n✅ {JSON_FILE} salvo · Mês: {mes_str}")

# ── GIT PUSH ─────────────────────────────────────────────────────────
msg = f"Operacional {mes_str} — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
for cmd in [['git','add',JSON_FILE],['git','commit','-m',msg],['git','push','origin','main']]:
    r = subprocess.run(cmd, capture_output=True, text=True)
    ok = r.returncode == 0 or 'nothing to commit' in r.stdout
    print(f"  {'✓' if ok else '⚠'} {' '.join(cmd[1:2])}")

print("\n🎉 Concluído!")
