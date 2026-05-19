#!/usr/bin/env python3
# ─────────────────────────────────────────────────────────────────────
# gerar_dados.py  —  Converte a planilha Excel em dados.json
# e faz o commit + push automático para o GitHub.
#
# USO:
#   python gerar_dados.py                        (usa template_dashboard_ip_2026.xlsx)
#   python gerar_dados.py minha_planilha.xlsx    (usa outro arquivo)
# ─────────────────────────────────────────────────────────────────────
import sys, os, json, re, subprocess
from datetime import datetime

try:
    import openpyxl
except ImportError:
    subprocess.run([sys.executable,"-m","pip","install","openpyxl","--break-system-packages","-q"])
    import openpyxl

# ── CONFIG ────────────────────────────────────────────────────────────
XLSX_FILE = sys.argv[1] if len(sys.argv) > 1 else "template_dashboard_ip_2026.xlsx"
JSON_FILE = "dados.json"
MES_PAT   = re.compile(r'^[a-záàâãéêíóôõúç]{2,4}/\d{2,4}$', re.IGNORECASE)

DIAS_MES = {
    "jan":31,"fev":28,"mar":31,"abr":30,"mai":31,"jun":30,
    "jul":31,"ago":31,"set":30,"out":31,"nov":30,"dez":31
}

# ── LEITURA ───────────────────────────────────────────────────────────
print(f"\n📊  Lendo: {XLSX_FILE}")
if not os.path.exists(XLSX_FILE):
    print(f"❌  Arquivo não encontrado: {XLSX_FILE}"); sys.exit(1)

wb  = openpyxl.load_workbook(XLSX_FILE, data_only=True)
wsn = next((s for s in wb.sheetnames if 'DADOS' in s.upper()), wb.sheetnames[0])
ws  = wb[wsn]
rows = [[str(c.value or '').strip() for c in row] for row in ws.iter_rows()]

# Encontrar linha de dados
data_start = -1
for i, row in enumerate(rows):
    if MES_PAT.match(row[0]):
        data_start = i; break

if data_start < 0:
    print("❌  Não encontrei dados de mês. Verifique se a planilha foi preenchida."); sys.exit(1)

# Mapear colunas (varre linhas de cabeçalho acima dos dados)
KEYS = {
    'CMM':   ['cmm'],
    'FM':    ['fm','fator'],
    'Nf':    ['nf','falhas g'],
    'Ni':    ['ni','interv'],
    'Pt':    ['pt','disposit'],
    'NCEi':  ['ncei','emerg. id','emerg id'],
    'NCEa':  ['ncea','emerg. at','emerg at'],
    'NCNEi': ['ncnei','n.em. id','n.emerg. id','não emerg'],
    'NCNEa': ['ncnea','n.em. at','n.emerg. at'],
    'Dm':    ['dm','dias'],
    'Ti':    ['ti','inoper'],
    'pOtimo':['% ótimo','%ótimo','ótimo','otimo'],
    'pBom':  ['% bom','%bom','bom'],
    'pReg':  ['% regular','%regular','regular'],
    'pRuim': ['% ruim','%ruim','ruim'],
    'lig':   ['total lig','ligações','ligacoes'],
    'nOp':   ['não opinou','nao opinou'],
}
found = {}
for ri in range(data_start):
    row_low = [c.lower().replace('\n',' ') for c in rows[ri]]
    for field, needles in KEYS.items():
        if field in found: continue
        for ci, cell in enumerate(row_low):
            if any(n in cell for n in needles):
                found[field] = ci; break

def num(row, ci):
    if ci is None: return 0.0
    try: return float(str(row[ci]).replace(',','.').strip() or 0)
    except: return 0.0

# Extrair dados por mês
meses = {}
for i in range(data_start, len(rows)):
    row = rows[i]
    mes = row[0].strip()
    if not MES_PAT.match(mes): continue
    d = {
        'mes':   mes,
        'CMM':   num(row, found.get('CMM')),
        'FM':    num(row, found.get('FM')),
        'Nf':    num(row, found.get('Nf')),
        'Ni':    num(row, found.get('Ni')),
        'Pt':    num(row, found.get('Pt')),
        'NCEi':  num(row, found.get('NCEi')),
        'NCEa':  num(row, found.get('NCEa')),
        'NCNEi': num(row, found.get('NCNEi')),
        'NCNEa': num(row, found.get('NCNEa')),
        'Dm':    num(row, found.get('Dm')) or DIAS_MES.get(mes[:3].lower(), 30),
        'Ti':    num(row, found.get('Ti')),
        'pOtimo':num(row, found.get('pOtimo')),
        'pBom':  num(row, found.get('pBom')),
        'pReg':  num(row, found.get('pReg')),
        'pRuim': num(row, found.get('pRuim')),
        'lig':   num(row, found.get('lig')),
        'nOp':   num(row, found.get('nOp')),
    }
    if d['CMM'] > 0 or d['Nf'] > 0 or d['NCEi'] > 0 or d['pOtimo'] > 0:
        meses[mes] = d

if not meses:
    print("❌  Planilha lida mas sem dados preenchidos. Preencha pelo menos um mês."); sys.exit(1)

print(f"✅  {len(meses)} mês(es) encontrado(s): {', '.join(meses.keys())}")

# ── SALVAR JSON ───────────────────────────────────────────────────────
payload = {
    "ultima_atualizacao": datetime.now().strftime('%d/%m/%Y %H:%M'),
    "meses": meses
}
with open(JSON_FILE, 'w', encoding='utf-8') as f:
    json.dump(payload, f, ensure_ascii=False, indent=2)
print(f"✅  {JSON_FILE} salvo com sucesso")

# ── GIT COMMIT + PUSH ─────────────────────────────────────────────────
msg = f"Dados atualizados — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
cmds = [
    ['git','add', JSON_FILE],
    ['git','commit','-m', msg],
    ['git','push','origin','main'],
]
print("\n📤  Enviando para o GitHub...")
for cmd in cmds:
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0 and 'nothing to commit' not in result.stdout:
        print(f"   ⚠  {' '.join(cmd)}: {result.stderr.strip()}")
    else:
        print(f"   ✓  {' '.join(cmd[1:2])}")

print("\n🎉  Concluído! O dashboard será atualizado em ~1 minuto.")
print(f"     Última atualização: {payload['ultima_atualizacao']}")
